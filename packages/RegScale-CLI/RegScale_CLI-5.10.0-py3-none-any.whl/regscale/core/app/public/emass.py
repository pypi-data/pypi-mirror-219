#!/usr/bin/env python3
# -*- coding: utf-8 -*-
""" eMASS integration to the CLI to allow support for eMASS documents """

# standard python imports
import os
from pathlib import Path

import click
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    check_file_path,
    create_progress_object,
    error_and_exit,
    get_current_datetime,
    get_file_type,
    reformat_str_date,
)
from regscale.models import regscale_id

SKIP_ROWS: int = 7
COLUMNS = ["L", "M", "N", "O"]


@click.group()
def emass():
    """[BETA] Performs bulk processing of eMASS files (Upload trusted data only)."""


@emass.command("get_template")
def get_template():
    """
    Fetch a template for the eMASS controls document
    """
    fetch_template_from_blob()


@emass.command("populate_controls")
@click.option(
    "--file_name",
    type=click.Path(exists=True, dir_okay=False, file_okay=True),
    required=True,
    prompt="Enter the full file path of the eMASS controls document.",
    help="Enter the full file path of the eMASS controls document to populate with RegScale data.",
)
@regscale_id(help="Enter the desired SSP ID # from RegScale.")
def populate_workbook(file_name: click.Path, regscale_id: int) -> None:
    """
    [BETA] Populate controls from a System Security Plan in RegScale into an eMASS formatted excel workbook.
    """
    populate_emass_workbook(file_name=file_name, regscale_id=regscale_id)


def fetch_template_from_blob() -> None:
    """
    Fetch a template for the eMASS controls document
    :return: None
    """
    logger = create_logger()
    app = Application()
    api = Api(app)

    # check if the artifacts folder exists
    check_file_path("artifacts")

    # get the template from the API
    template = api.get(
        url="https://regscaleblob.blob.core.windows.net/blob/eMASS_Control_Template.xlsx",
        headers={},
    )

    # write the template to a file
    with open(f".{os.sep}artifacts{os.sep}eMASS_Template.xlsx", "wb") as f:
        f.write(template.content)
    logger.info(f"Template saved to .{os.sep}artifacts{os.sep}eMASS_Template.xlsx")


def populate_emass_workbook(file_name: Path, regscale_id: int) -> None:
    """
    Function to populate an eMASS workbook with control assessments from RegScale
    :param Path file_name: Path to the eMASS control workbook
    :param int regscale_id: ID of the SSP in RegScale to get the controls & assessments from
    :return: None
    """
    logger = create_logger()
    # make sure the user gave a path to an Excel workbook
    if get_file_type(file_name) not in [".xlsx", ".xls"]:
        error_and_exit(
            "Please provide a file path to an Excel workbook in .xlsx or .xls format."
        )

    # convert file_name to a Path object
    file_name = Path(file_name)

    # initialize the Application and API classes
    app = Application()
    api = Api(app)

    # populate the controls in the Excel workbook
    output_name = populate_controls(file_name=file_name, ssp_id=regscale_id, api=api)
    logger.info(
        "Please open %s and verify the data before uploading into eMASS.", output_name
    )


def format_controls(file_data_dict: dict, file_name: str) -> tuple[dict, dict]:
    """
    Function to format controls to your desired format
    :param dict file_data_dict: Dictionary of an Excel file column
    :param str file_name: Name of the file file_data_dict is from
    :return: tuple of two dicts: raw_controls and formatted_controls
    :rtype: tuple[dict, dict]
    """
    # convert the control names to match RegScale control names
    raw_controls = ({},)
    formatted_controls = {}
    try:
        raw_controls = {
            val.lower(): key for key, val in file_data_dict["Control Acronym"].items()
        }
        # Remove duplicate values in dictionary
        raw_controls = dict(raw_controls)

        # create list of formatted controls
        formatted_controls = [
            v.lower().replace("(", ".").replace(")", "")
            for v in file_data_dict["AP Acronym"].values()
        ]
    except KeyError:
        error_and_exit(
            f"{file_name} doesn't match the expected eMASS format.\nPlease view an example template here: https://regscale.readme.io/docs/emass-beta#template"
        )

    return raw_controls, formatted_controls


def populate_controls(file_name: Path, ssp_id: int, api: Api) -> Path:
    """
    Populate controls from a System Security Plan in RegScale into an eMASS formatted excel workbook
    :param Path file_name: path to the Excel workbook to populate with controls from SSP
    :param int ssp_id: ID for a System Security Plan from RegScale
    :param Api api: API Object
    :return: Path to output file
    :rtype: Path
    """
    job_progress = create_progress_object()
    logger = create_logger()
    # create the GraphQL query
    query = f"""
        query {{
          controls:controlImplementations(
            take: 50
            skip: 0
            where: {{
              parentId: {{ eq: {ssp_id} }}
          parentModule: {{ eq: "securityplans" }}
          assessments: {{ any: true }}
        }}
      ) {{
        items {{
          id
          control {{
            controlId
          }}
          assessments {{
            id
            actualFinish
            assessmentResult
            summaryOfResults
            leadAssessor {{
              firstName
              lastName
            }}
          }}
        }}
        totalCount
        pageInfo {{
          hasNextPage
        }}
      }}
    }}
    """

    # get the data from GraphQL
    response = api.graph(query=query)

    # try to get the items from the GraphQL response
    try:
        controls = response["controls"]["items"]
    except KeyError:
        controls = []

    total_controls = api.get(
        f"{api.config['domain']}/api/controlImplementation/getCountByParent/{ssp_id}/securityplans"
    )

    if not total_controls.ok:
        error_and_exit(
            f"Received unexpected response: {total_controls.status_code}\n{total_controls.text}",
        )

    if len(controls) > 0:
        logger.info(
            "Received %s/%s controls with Assessments. Total control count for SSP #%s in RegScale: %s.",
            len(controls),
            response["controls"]["totalCount"],
            ssp_id,
            total_controls.text,
        )
    else:
        error_and_exit(
            "The RegScale SSP provided has no assessments associated with the controls. "
            + "Please add assessments to the controls and try again."
        )

    # load the Excel file in pandas to find row # to update the data
    file_data = pd.read_excel(file_name, skiprows=SKIP_ROWS - 2)

    # load the workbook using openpyxl to retain worksheet styling
    wb = load_workbook(file_name)

    # set the sheet to the first sheet in the provided workbook
    sheet = wb.active

    # convert to a dictionary
    file_data_dict = file_data.to_dict()

    # format the controls
    raw_controls, formatted_controls = format_controls(
        file_data_dict=file_data_dict, file_name=file_name.name
    )

    # create variable to count number of rows updated and skipped
    update_counter: int = 0
    skipped_counter: int = 0

    # create a list of all the control ids from the GraphQL query
    regscale_control_ids = [ctrl["control"]["controlId"] for ctrl in controls]

    # create comment & fill attribute for columns with missing data
    comment = Comment(
        text=f"SSP #{ssp_id} doesn't contain an assessment associated with this control.",
        author="RegScale CLI",
        height=150,
    )
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    with job_progress:
        populating_controls = job_progress.add_task(
            f"[#21a5bb]Analyzing controls in {file_name.name}...",
            total=len(formatted_controls),
        )
        # iterate through the controls of the provided Excel workbook
        for control in formatted_controls:
            # get the row number for the current control
            row_number = formatted_controls.index(control) + SKIP_ROWS

            # see if the current formatted control has an assessment by
            # comparing it to our GraphQL query results
            if control in regscale_control_ids:
                # get the control from the GraphQL results
                regscale_control = controls[regscale_control_ids.index(control)]
                # get the assessment for the control
                assessment = regscale_control["assessments"][0]

                # determine compliance status
                if assessment["assessmentResult"] == "Pass":
                    compliance_status = "Compliant"
                elif assessment["assessmentResult"] in ["Fail", "Partial Pass"]:
                    compliance_status = "Non-Compliant"
                else:
                    compliance_status = "Not Applicable"

                # reformat the assessment actualFinish date if it is, populate it
                finish_date = (
                    reformat_str_date(assessment["actualFinish"])
                    if assessment["actualFinish"]
                    else None
                )

                # map the control to the Excel spreadsheet
                sheet[f"L{row_number}"] = compliance_status
                if finish_date:
                    sheet[f"M{row_number}"] = finish_date
                else:
                    sheet[f"M{row_number}"].comment = Comment(
                        text=f"Assessment #{assessment['id']} in RegScale doesn't have a finish date.",
                        author="RegScale CLI",
                        height=150,
                    )
                    sheet[f"M{row_number}"].fill = yellow_fill
                sheet[
                    f"N{row_number}"
                ] = f'{assessment["leadAssessor"]["firstName"]} {assessment["leadAssessor"]["lastName"]}'
                if assessment["summaryOfResults"]:
                    sheet[f"O{row_number}"] = assessment["summaryOfResults"]
                else:
                    sheet[f"O{row_number}"].comment = Comment(
                        text=f"Assessment #{assessment['id']} in RegScale doesn't have any information in Summary of Results.",
                        author="RegScale CLI",
                        height=150,
                    )
                    sheet[f"O{row_number}"].fill = yellow_fill

                # update the counter
                update_counter += 1
            else:
                # increment the skip counter
                skipped_counter += 1

                # highlight and add a comment
                for column in COLUMNS:
                    sheet[f"{column}{row_number}"].comment = comment
                    sheet[f"{column}{row_number}"].fill = yellow_fill
            # update the progress bar
            job_progress.update(populating_controls, advance=1)

    # add the date and time to the output filename
    output_name = Path(
        os.path.join(
            file_name.parent,
            file_name.stem
            + get_current_datetime("_Updated_%Y%m%d_%H%M%S")
            + file_name.suffix,
        )
    )

    # save the updated workbook
    wb.save(output_name)

    logger.info(
        "%s has been created with %i update(s). %i row(s) were skipped because of missing controls in SSP #%i.",
        output_name.name,
        update_counter,
        skipped_counter,
        ssp_id,
    )
    # return the output path
    return output_name
