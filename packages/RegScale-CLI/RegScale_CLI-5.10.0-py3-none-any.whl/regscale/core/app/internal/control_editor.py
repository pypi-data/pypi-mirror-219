#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Module to allow user to make changes to Control Implementations in an Excel spreadsheet for a user-friendly experience
"""

# standard python imports
import math
import os
import shutil
from pathlib import Path

import click
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation

from regscale.core.app.api import Api
from regscale.core.app.application import Application
from regscale.core.app.logz import create_logger
from regscale.core.app.utils.app_utils import (
    error_and_exit,
    get_current_datetime,
    check_file_path,
    check_empty_nan,
)
from regscale.models.app_models.click import regscale_id, regscale_module
from regscale.models.regscale_models.control_implementation import (
    ControlImplementation,
    Control,
)


@click.group(name="control_editor")
def control_editor():
    """
    Performs actions on Control Editor Feature to edit controls to RegScale.
    """


# Get data and pull into Excel worksheets.


@control_editor.command(name="generate")
@regscale_id()
@regscale_module()
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path for created excel files to be saved to.",
    default=Path("./artifacts"),
    required=True,
)
def generate_data_download(regscale_id: str, regscale_module: str, path: Path):
    """
    This function will build and populate a spreadsheet of all control implementations
    with the selected RegScale Parent Id and RegScale Module.
    """
    data_load(regscale_id=regscale_id, regscale_module=regscale_module, path=path)


def data_load(regscale_id: str, regscale_module: str, path: Path):
    """Function takes organizer record and module and build excel worksheet of control implementations.

    :param regscale_id: RegScale Parent Id
    :param regscale_module: RegScale Parent Module
    :param path: directory of file location
    """
    logger = create_logger()
    app = Application()
    api = Api(app)

    # Making directory for files

    check_file_path(path)

    workbook = Workbook()
    ws = workbook.active
    ws.title = f"Impls_PId({regscale_id}_{regscale_module})"
    workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))
    shutil.copy(
        os.path.join(path, "all_implementations.xlsx"),
        os.path.join(path, "old_implementations.xlsx"),
    )

    # Loading data from RegScale database into two workbooks.

    body = """
            query{
                controlImplementations (skip: 0, take: 50, where: {parentId: {eq: parent_id} parentModule: {eq: "parent_module"}}) {
                    items {
                        id
                        controlID
                        controlOwnerId
                        controlOwner {
                            firstName
                            lastName
                            userName
                        }
                        control {
                            title
                            description
                            controlId
                        }
                        status
                        policy
                        implementation
                        responsibility
                        inheritable
                        parentId
                        parentModule
                    }
                    totalCount
                    pageInfo {
                        hasNextPage
                    }
                }
            }""".replace(
        "parent_module", regscale_module
    ).replace(
        "parent_id", str(regscale_id)
    )

    existing_implementation_data = api.graph(query=body)

    if existing_implementation_data["controlImplementations"]["totalCount"] > 0:
        raw_data = existing_implementation_data["controlImplementations"]["items"]

        all_imps = []
        for item in raw_data:
            Id = item["id"]
            ControlId = item["controlID"]
            ControlOwnerId = item["controlOwnerId"]
            ControlOwner = (
                str(item["controlOwner"]["lastName"]).strip()
                + ", "
                + str(item["controlOwner"]["firstName"]).strip()
                + " ("
                + str(item["controlOwner"]["userName"]).strip()
                + ")"
            )
            ControlName = item["control"]["controlId"]
            ControlTitle = item["control"]["title"]
            Description = item["control"]["description"]
            Status = item["status"]
            Policy = item["policy"]
            Implementation = item["implementation"]
            Responsibility = item["responsibility"]
            Inheritable = item["inheritable"]

            all_imps.append(
                [
                    Id,
                    ControlId,
                    ControlOwnerId,
                    ControlOwner,
                    ControlName,
                    ControlTitle,
                    Description,
                    Status,
                    Policy,
                    Implementation,
                    Responsibility,
                    Inheritable,
                ]
            )

        all_imps_df = pd.DataFrame(
            all_imps,
            columns=[
                "Id",
                "ControlId",
                "ControlOwnerId",
                "ControlOwner",
                "ControlName",
                "ControlTitle",
                "Description",
                "Status",
                "Policy",
                "Implementation",
                "Responsibility",
                "Inheritable",
            ],
        )

        with pd.ExcelWriter(
            os.path.join(path, "all_implementations.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_imps_df.to_excel(
                writer,
                sheet_name=f"Impls_PId({regscale_id}_{regscale_module})",
                index=False,
            )

        with pd.ExcelWriter(
            os.path.join(path, "old_implementations.xlsx"),
            mode="a",
            engine="openpyxl",
            if_sheet_exists="overlay",
        ) as writer:
            all_imps_df.to_excel(
                writer,
                sheet_name=f"Impls_PId({regscale_id}_{regscale_module})",
                index=False,
            )
    else:
        error_and_exit(
            "No records exist for the given RegScale Id and RegScale Module."
        )

    # Adding Data validation to "old_implementations.xlsx" file that will be used as reference.

    workbook2 = load_workbook(os.path.join(path, "old_implementations.xlsx"))
    worksheet2 = workbook2.active
    worksheet2.protection.sheet = True
    workbook2.save(filename=os.path.join(path, "old_implementations.xlsx"))

    # Adding Data Validation to "all_implementations.xlsx" file to be adjusted internally by clients.

    workbook = load_workbook(os.path.join(path, "all_implementations.xlsx"))
    worksheet = workbook.active
    worksheet.protection.sheet = True

    dv1 = DataValidation(
        type="list",
        formula1='"Not Implemented, Fully Implemented, In Remediation, Not Applicable, Inherited, Planned"',
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv2 = DataValidation(
        type="list",
        formula1='"Provider, Customer, Shared, Not Applicable"',
        allow_blank=True,
        showDropDown=False,
        error="Your entry is not one of the available options",
        errorTitle="Invalid Entry",
        prompt="Please select from the list",
    )
    dv3 = DataValidation(type="list", formula1='"TRUE, FALSE"', allow_blank=True)

    worksheet.add_data_validation(dv1)
    worksheet.add_data_validation(dv2)
    worksheet.add_data_validation(dv3)
    dv1.add("H2:H1048576")
    dv2.add("K2:K1048576")
    dv3.add("L2:L1048576")

    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

        if col == ["H", "I", "J", "K", "L"]:  # columns to edit
            for cell in worksheet[col]:
                cell.protection = Protection(locked=False)

    workbook.save(filename=os.path.join(path, "all_implementations.xlsx"))

    logger.info("Successfully created the directory %s.", path)
    logger.info("All files are located within directory.")

    logger.info(
        "Your data has been loaded into your excel workbook. Please open the all_implementations workbook and make your desired changes."
    )
    return None


# Save Spreadsheet if file changed, append Update API changes that were manually entered in an Excel worksheet


@control_editor.command(name="load")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path where excel workbooks are located.",
    default=Path("./artifacts"),
    required=True,
)
@click.option(
    "--skip_prompt",
    type=click.BOOL,
    help="To Skip (Y/N) Prompt, input True.",
    default=False,
    required=False,
)
def generate_db_update(path: Path, skip_prompt: bool):
    """
    This function will check changes made to spreadsheet and upload any changes made to RegScale.

    """
    db_update(path, skip_prompt)


def db_update(path: Path, skip_prompt=True):
    """Function will check changes made by user and upload any changes to RegScale.

    :param path: directory of file location
    :param skip_prompt: boolean to skip prompt save message, defaults to True
    :return: None
    """
    logger = create_logger()

    logger.info(
        "Proceed only after you have made the necessary changes and have saved file."
    )

    x = "y" if skip_prompt else input("Ready to Proceed (Y/N): ").lower()

    if x[0] == "y":
        df = load_workbook(os.path.join(path, "all_implementations.xlsx"))

        sheet_name = df.sheetnames[0]
        sheet_name = sheet_name[sheet_name.find("(") + 1 : sheet_name.find(")")].split(
            "_"
        )
        regscale_parent_id, regscale_module = set(sheet_name)

        df1 = pd.read_excel(
            os.path.join(path, "all_implementations.xlsx"), sheet_name=0, index_col="Id"
        )

        df2 = pd.read_excel(
            os.path.join(path, "old_implementations.xlsx"), sheet_name=0, index_col="Id"
        )

        if df1.equals(df2):
            error_and_exit("No differences detected.")

        else:
            logger.warning("*** WARNING *** Differences Found.")

            # Logs changes to txt file

            diff_mask = (df1 != df2) & ~(df1.isnull() & df2.isnull())
            ne_stacked = diff_mask.stack()
            changed = ne_stacked[ne_stacked]
            changed.index.names = ["Id", "Column"]
            difference_locations = np.where(diff_mask)
            changed_to = df1.values[difference_locations]
            changed_from = df2.values[difference_locations]
            changes = pd.DataFrame(
                {"From": changed_from, "To": changed_to}, index=changed.index
            )
            changes.to_csv(
                os.path.join(path, "differences.txt"),
                header=True,
                index=True,
                sep=" ",
                mode="a",
            )

            upload_data(regscale_parent_id, regscale_module, path)

    logger.info(
        "Please check differences.txt file located in artifacts folder to see changes made."
    )
    return None


def upload_data(regscale_parent_id: int, regscale_module: str, path: Path) -> None:
    """
    Batch uploads updated control implementation statements to the provided RegScale parent ID.
    :param int regscale_parent_id: RegScale parent ID
    :param str regscale_module: RegScale parent module
    :param Path path: file path where control spreadsheet resides
    :raises: requests.exceptions.RequestException if API call encountered an error
    :return: None
    """
    app = Application()
    config = app.config
    api = Api(app)

    diff = pd.read_csv(
        os.path.join(path, "differences.txt"), header=0, sep=" ", index_col=None
    )
    ids = []
    for i, row in diff.iterrows():
        ids.append(row["Id"])

    id_df = pd.DataFrame(ids, index=None, columns=["Id"])
    id_df2 = id_df.drop_duplicates()

    reader = pd.read_excel(os.path.join(path, "all_implementations.xlsx"))
    updates = reader[reader["Id"].isin(id_df2["Id"])]
    updates = reader.T.to_dict()
    updated_implementations = [
        ControlImplementation(
            id=i["Id"],
            controlOwnerId=i["ControlOwnerId"],
            control=Control(
                title=i["ControlTitle"],
                description=i["Description"],
                controlId=i["ControlName"],
            ).dict(),
            status=i["Status"],
            implementation=check_empty_nan(i["Implementation"]),
            policy=check_empty_nan(i["Policy"]),
            controlID=i["ControlId"],
            responsibility=check_empty_nan(i["Responsibility"]),
            parentId=regscale_parent_id,
            parentModule=regscale_module,
            inheritable=check_inheritable(i["Inheritable"]),
            lastUpdatedById=app.config["userId"],
            dateLastUpdated=get_current_datetime(),
        ).dict()
        for i in updates.values()
    ]

    api.update_server(
        url=config["domain"] + "/api/controlImplementation",
        json_list=updated_implementations,
        message="Working on uploading updated control implementations to RegScale.",
        config=config,
        method="put",
    )


# Delete and remove files from user's system.


@control_editor.command(name="delete_files")
@click.option(
    "--path",
    type=click.Path(exists=False, dir_okay=True, path_type=Path),
    help="Provide the desired path of file location.",
    default=Path("./artifacts"),
    required=True,
)
def generate_delete_file(path: Path):
    """This command will delete files used during the Control editing process."""
    delete_file(path)


def delete_file(path: Path):
    """
    Deletes files used during the process.

    :param path: directory of file location
    :return: None
    """
    logger = create_logger()

    os.remove(os.path.join(path, "all_implementations.xlsx"))
    os.remove(os.path.join(path, "old_implementations.xlsx"))
    if os.path.isfile(os.path.join(path, "differences.txt")):
        os.remove(os.path.join(path, "differences.txt"))
    else:
        pass
    os.rmdir(path)
    logger.info("Files have been deleted. Thank you.")
    return None


def check_inheritable(
    value,
):  # this function has to be checked separate to account for API only accpeting False Boolean unlike other class paramas
    """This function takes a given value for an inheritable and checks if value is empty or NaN based on value type.
    :param value: A string or float object
    :return: A string value, float value. or False
    :rtype: str, float, or False
    """
    if isinstance(value, str) and value.strip() == "":
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    return value
