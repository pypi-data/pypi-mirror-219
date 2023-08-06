from threemystic_cloud_cmdb.cloud_providers.general.cmdb_connector.base_class.base import cloud_cmdb_general_cmdb_connector_base as base
import urllib
from time import sleep
from openpyxl.utils import get_column_letter
from math import floor
from decimal import Decimal
from requests.exceptions import HTTPError

# References
# https://learn.microsoft.com/en-us/graph/api/overview?view=graph-rest-1.0
# https://learn.microsoft.com/en-us/graph/query-parameters?tabs=http
class cloud_cmdb_general_cmdb_connector_ms365(base):
  def __init__(self, *args, **kwargs):
    super().__init__(logger_name= "cloud_cmdb_general_cmdb_connector_ms365", *args, **kwargs)
  

  
  def get_cloud_share(self, *args, **kwargs):
    return "ms365"
  
  def __close_session(self, *args, **kwargs):
    if self.get_cmdb_file() is None:
      return
    if self.get_cmdb_file().get('id') is None:
      return
    
    self._get_ms_graph().close_session(session_config = {
      "type":"workbook",
      "graph_resource": self._get_ms_graph_resource(),
      "graph_resource_id": self._get_ms_graph_resource_id(),
      "drive_id": self.get_cmdb_file().get('id'),
      "persist_changes": True
    })

    self._get_ms_graph().close_session(session_config = {
      "type":"workbook",
      "graph_resource": self._get_ms_graph_resource(),
      "graph_resource_id": self._get_ms_graph_resource_id(),
      "drive_id": self.get_cmdb_file().get('id'),
      "persist_changes": False
    })
    
  def _validate_cmdb_init(self, *args, **kwargs):
    self._validate_cmdb_file()
    self._validate_workbook_worksheets()
    self._validate_workbook_worksheets_tables()
    self._validate_workbook_worksheets_tables_columns()   

    self.__close_session()
    

  
  def get_existing_columns_sorted_by_index(self, *args, **kwargs):    
    if hasattr(self, "_ms365_existing_columns_sorted_by_index"):
      return self._ms365_existing_columns_sorted_by_index
    
    self._ms365_existing_columns_sorted_by_index = {}
    for container_key, table_data in self._get_worksheet_table_data().items():
      self._ms365_existing_columns_sorted_by_index[container_key] = [
        sorted_column.get("name") for sorted_column in sorted(list(table_data.get("extra_columns").get("value").values()), key=lambda x: x.get("index"), reverse= False)
      ]
    
    return self.get_existing_columns_sorted_by_index()

  
  def _get_workbook_table_name(self, sheet_name, *args, **kwargs):
    return f'cmdb_{self.get_cmdb_data_containers_display_key()[sheet_name]}'

  def get_cmdb_name(self, *args, **kwargs):
    if hasattr(self, "_cmdb_name_ext"):
      return self._cmdb_name_ext
    
    if not super().get_cmdb_name().endswith(".xlsx"):
      self._cmdb_name_ext = self.get_common().helper_type().string().set_case(
        string_value= f"{super().get_cmdb_name()}.xlsx",
        case= "lower"
      )
    
    return self.get_cmdb_name(*args, **kwargs)
  
  def _get_ms_graph_drive_id(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph_drive_id"):
      return self._ms365_graph_drive_id
    
    self._ms365_graph_drive_id = self.get_cloud_share_config_value(
      config_key= self.get_cloud_share()
    )["drive_id"]
    return self._get_ms_graph_drive_id(*args, **kwargs)

  def _get_ms_graph_resource_id(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph_resource_id"):
      return self._ms365_graph_resource_id
    
    self._ms365_graph_resource_id = self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') if self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') != "me" else None
    return self._get_ms_graph_resource_id(*args, **kwargs)

  def _get_ms_graph_resource(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph_resource"):
      return self._ms365_graph_resource
    
    self._ms365_graph_resource = "me" if self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') == "me" else f"groups"
    return self._get_ms_graph_resource(*args, **kwargs)

  def _get_ms_graph(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph"):
      return self._ms365_graph
    
    self._ms365_graph = self.get_common().graph().graph(
      graph_method= "msgraph", 
      credentials= self.get_cloud_client().get_tenant_credential(
        tenant= self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('tenant_id')))
    
    return self._get_ms_graph(*args, **kwargs)
   
  def get_cmdb_file(self, *args, **kwargs):
    if not hasattr(self, "_ms36_cmdb_file"):
      return None
    
    return self._ms36_cmdb_file
  
  def __set_cmdb_file(self, file_details, *args, **kwargs):
    self._ms36_cmdb_file = file_details

  def _get_ms_graph_base_path(self, drive_item_id, *args, **kwargs):
    return f"{drive_item_id}" if self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') == "me" else f"items/{drive_item_id}"
    
  def _validate_cmdb_file(self, *args, **kwargs):

    try:
      local_drive_options = self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self._get_ms_graph_drive_id()[-1].get('id') )}/children?$select=id,name,file")
      )
    except Exception as err:
      self.__close_session()
      raise err

    if local_drive_options.get("value") != None:
      for item in local_drive_options.get("value"):
        if self.get_common().helper_type().string().set_case(
          string_value= item.get("name"),
          case= "lower"
        ) != self.get_cmdb_name():
          continue

        if item.get("file") is None:
          continue
        

        self.__set_cmdb_file(file_details= item)
        break
    
    if self.get_cmdb_file() is not None:
      return
    
    self._create_cmdb_file()
    self._validate_cmdb_file()

  def _create_cmdb_file(self, *args, **kwargs):
    
    from tempfile import NamedTemporaryFile
    from openpyxl import Workbook
    
    excel_doc = Workbook()
    while len(excel_doc.sheetnames) > 0:
      excel_doc.remove(excel_doc[excel_doc.sheetnames[0]])
    
    for sheet_key, sheet_name in self.get_cmdb_data_containers_key_display().items():
      excel_sheet = excel_doc.create_sheet(sheet_name)
      excel_sheet.append(self.get_cmdb_data_containers_columns().get(sheet_key))
      excel_sheet.freeze_panes = "A2"
    
    
    with NamedTemporaryFile() as tmp:

      excel_doc.save(tmp.name)
      tmp.seek(0)

      try:
        self._get_ms_graph().send_request(
          url = self._get_ms_graph().generate_graph_url(
            resource= self._get_ms_graph_resource(), 
            resource_id= self._get_ms_graph_resource_id(), 
            base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self._get_ms_graph_drive_id()[-1].get('id') )}:/{urllib.parse.quote(self.get_cmdb_name())}:/content"),
          data = tmp.read(),
          params= {"@microsoft.graph.conflictBehavior": "replace"},
          headers= {"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
          method= "put"
        )
        
      except Exception as err:
        self.__close_session()
        raise err
      

  
  def _get_worksheet_data(self, *args, **kwargs):
    if hasattr(self, "_worksheet_data"):
      return self._worksheet_data
    
    return None

  def _validate_workbook_worksheets(self, *args, **kwargs):

    try:
      worksheets_response = self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        method= "get"
      )
    except Exception as err:
      self.__close_session()
      raise err

    self._worksheet_data = None
    if worksheets_response.get("value") is None:
      for data_container_key, data_container_display in self.get_cmdb_data_containers_key_display().items():
        self._add_workbook_worksheet(sheet_key= data_container_key, sheet_name= data_container_display)
      return self._validate_workbook_worksheets(*args, **kwargs)
    
    if len(worksheets_response.get("value")) < 1:
      for data_container_key, data_container_display in self.get_cmdb_data_containers_key_display().items():
        self._add_workbook_worksheet(sheet_key= data_container_key, sheet_name= data_container_display)
      return self._validate_workbook_worksheets(*args, **kwargs)
    
    self._worksheet_data = {self.get_cmdb_data_containers_display_key()[worksheet["name"]]:worksheet for worksheet in worksheets_response["value"] if self.get_cmdb_data_containers_display_key().get(worksheet["name"]) is not None}
    missing_worksheets = False
    for data_container_key, data_container_display in self.get_cmdb_data_containers_key_display().items():
      if self._worksheet_data.get(data_container_key) is not None:
        continue

      self._add_workbook_worksheet(sheet_key= data_container_key,sheet_name= data_container_display)
      missing_worksheets = True
    
    if missing_worksheets:      
      return self._validate_workbook_worksheets(*args, **kwargs)
    
    return None

  def _add_workbook_worksheet(self, sheet_name, sheet_key = None, *args, **kwargs):
    
    try:
      response = (self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        data = {
        "name": sheet_name
        },
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "post"
      ))
    except Exception as err:
      self.__close_session()
      raise err

    self.init_workbook_worksheet(sheet_key= sheet_key, sheet_name= sheet_name, sheet_id= response.get("id"))
  
  def init_workbook_worksheet(self, sheet_name, sheet_key, sheet_id = None, *args, **kwargs):
    
    try:
      if self.get_common().helper_type().string().is_null_or_whitespace(string_value= sheet_id):
        sheet_id= self._get_worksheet_data()[sheet_key].get('id') if self._get_worksheet_data().get(sheet_key) is not None else sheet_name
      
      return self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{sheet_id}/range(address='A1:{get_column_letter(len(self.get_cmdb_data_containers_columns().get(sheet_key)))}1')"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        data= {
          "values": [self.get_cmdb_data_containers_columns().get(sheet_key)]
        },
        method= "patch"
      )
    except Exception as err:
      self.__close_session()
      raise err
  
  def _get_workbook_worksheets_table_columns(self, sheet_key, table_id, *args, **kwargs):
    
    columns = None
    max_attempt = 5
    for attempt in (range(max_attempt)):
      try:
        columns = self._get_ms_graph().send_request(
          url = self._get_ms_graph().generate_graph_url(
            resource= self._get_ms_graph_resource(), 
            resource_id= self._get_ms_graph_resource_id(), 
            base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{table_id}/columns?$select=id,index,name"),
          method= "get"
        )
      except HTTPError as err:
        if err.response.status_code == 404:
          self.get_common().helper_type().requests().expodential_backoff_wait(attempt= attempt, auto_sleep= True)
          self.__close_session()
          continue
        self.__close_session()
        raise err
      except Exception as err:
        self.__close_session()
        raise err

    
    if columns is None:
      self.__close_session()
      return {}
    
    if columns.get("value") is None:
      return []
    
    columns["value"] = {
          column.get("name"):column
          for column in columns["value"]}
    return columns

  def _validate_workbook_worksheets_table(self, sheet_key, sheet_name, table_response, *args, **kwargs):
    if table_response is None:
      return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)

    if table_response.get("value") is None:
      return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)

    if len(table_response.get("value")) < 1:
      return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)
    
    for table in table_response.get("value"):
      if table.get("name") == self._get_workbook_table_name(sheet_name= sheet_name):
        table["extra_columns"] = self._get_workbook_worksheets_table_columns(
          sheet_key= sheet_key,
          table_id= table.get('id')
        )
          
        return table
      
    return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)

  def _get_worksheet_table_data(self, *args, **kwargs):
    if hasattr(self, "_worksheet_table_data"):
      return self._worksheet_table_data
    
    return None
  
  def _validate_workbook_worksheets_tables(self, *args, **kwargs):
    self._worksheet_table_data = {}
    for sheet_key, sheet_name in self.get_cmdb_data_containers_key_display().items():
      self._worksheet_table_data[sheet_key] = self._validate_workbook_worksheets_table(sheet_key= sheet_key, sheet_name= sheet_name, table_response= self._get_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name))

  def _get_workbook_worksheet_table(self, sheet_key, sheet_name, *args, **kwargs):
    try:
      return self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        method= "get"
      )
    except Exception as err:
      self.__close_session()
      raise err

  def _add_workbook_worksheet_table(self, sheet_key, sheet_name, *args, **kwargs):
    try:
      response = self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/add"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        data = {
        "address": f"{sheet_name}!A1:{get_column_letter(len(self.get_cmdb_data_containers_columns().get(sheet_key)))}1",
        "hasHeaders": True
        },
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "post"
      )
    except Exception as err:
      self.__close_session()
      raise err

    table_id = response.pop("id")
    response["name"] = self._get_workbook_table_name(sheet_name= sheet_name)
    
    try:
      return self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{table_id}"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        data = response,
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "patch"
      )
    except Exception as err:
      self.__close_session()
      raise err
    
    # hoping to ensure everything is pushed up
    self.__close_session()

  def _validate_workbook_worksheets_tables_columns(self, *args, **kwargs):
    for sheet_key in self.get_cmdb_data_containers_key_display().keys():
      self._validate_workbook_worksheets_tables_column(sheet_key= sheet_key)
  
  def _add_workbook_worksheets_tables_column(self, sheet_key, column, index, *args, **kwargs):
    try:
      return self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{self._get_worksheet_table_data()[sheet_key].get('id')}/columns/add"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True
        },
        data = {
          "index": index,
          "name": column
        },
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "post"
      )
    except Exception as err:
      self.__close_session()
      raise err

  def _validate_workbook_worksheets_tables_column(self, sheet_key, *args, **kwargs):
    column_index = -1
    add_columns = []
    
    if self._get_worksheet_table_data()[sheet_key].get("extra_columns") is None:
      self._get_worksheet_table_data()[sheet_key]["extra_columns"] = self._get_workbook_worksheets_table_columns(
          sheet_key= sheet_key,
          table_id= self._get_worksheet_table_data()[sheet_key].get('id')
        )
    
    for column in self.get_cmdb_data_containers_columns_raw_display_byid()[sheet_key].keys():
      column_index += 1

      if column not in self._get_worksheet_table_data()[sheet_key].get("extra_columns").get("value"):        
        add_columns.append({
          "index": column_index,
          "previous_index": column_index - 1,
          "next_index": column_index + 1,
          "column": column
        })
    if len(add_columns) < 1:
      return
  
    for column in add_columns:
      self._add_workbook_worksheets_tables_column(
        sheet_key=sheet_key,
        column= column.get("column"),
        index= column["index"]
      )
    
    self._worksheet_table_data[sheet_key] = self._validate_workbook_worksheets_table(
      sheet_key= sheet_key, sheet_name= self.get_cmdb_data_containers_key_display()[sheet_key], 
      table_response= self._get_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= self.get_cmdb_data_containers_key_display()[sheet_key]))


  def _sync_data(self, *args, **kwargs):
    try:
      self.__process_sync_data(*args, **kwargs)
    finally:
      self.__close_session()
      
  def __process_sync_data(self, *args, **kwargs):
    
    for sheet_key, processed_data in self._processed_report_data.items():
      existing_data = self.__sync_data_get_existing(sheet_key= sheet_key)
      if len(existing_data) < 1:
        self.__sync_data_add_data(sheet_key= sheet_key, insert_data= processed_data)
        continue
      
      cmdb_id_index = self.get_existing_columns_sorted_by_index()[sheet_key].index(self.get_cmdb_data_containers_columns_raw_byid_display()[sheet_key]["cmdb_id"])
      delete_index= -1
      if self.get_cmdb_data_containers_columns_raw_byid_display()[sheet_key].get("deleted") is not None:
        delete_index = self.get_existing_columns_sorted_by_index()[sheet_key].index(self.get_cmdb_data_containers_columns_raw_byid_display()[sheet_key]["deleted"])

      insert_data = []
      update_data = []

      while len(processed_data) > 0:
        row_cmdb_id = self.get_common().helper_type().string().set_case(
          string_value= processed_data[-1][cmdb_id_index],
          case= "lower"
        )
        if row_cmdb_id not in existing_data:
          insert_data.append(processed_data.pop())
          continue

        check_row = processed_data.pop()
        existing_row = existing_data.pop(row_cmdb_id)

        has_update_data = False
        for col_key, column in self.get_cmdb_data_containers_columns_raw_byid_display()[sheet_key].items():
          col_index = self.get_existing_columns_sorted_by_index()[sheet_key].index(column)
          
          if not self._get_cmdb_data_containers_validation(
              container= self.get_cmdb_data_containers_columns_raw()[sheet_key][col_key]
            )({
              "existing": existing_row.get("values")[0][col_index],
              "new": check_row[col_index]
            }
            ):
            
            if(self.get_common().helper_type().string().is_null_or_whitespace(string_value= existing_row.get("values")[0][col_index]) and 
              self.get_common().helper_type().string().is_null_or_whitespace(string_value= check_row[col_index])):
              continue

            existing_row.get("values")[0][col_index] = check_row[col_index]
            has_update_data = True
        
        if delete_index > -1:
          if not self.get_common().helper_type().string().is_null_or_whitespace(string_value= existing_row.get("values")[0][delete_index]):
            existing_row.get("values")[0][delete_index] = None
            has_update_data = True
            
        if has_update_data:
          update_data.append(existing_row)
      
      if delete_index > -1:
        while len(existing_data) > 0:
          _, deleted_data = existing_data.popitem()
          
          if not self.get_common().helper_type().string().is_null_or_whitespace(string_value= deleted_data.get("values")[0][delete_index]):
            continue

          deleted_data.get("values")[0][delete_index] = self._get_delete_time()
          update_data.append(deleted_data)
      
      self.__sync_data_update_data(sheet_key= sheet_key, update_data= update_data)
      self.__sync_data_add_data(sheet_key= sheet_key, insert_data= insert_data)

  def __sync_data_update_data_process(self, sheet_key, update_data, *args, **kwargs):

    return (self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= "$batch",
        resource_id= None,
        base_path= None),
      data = {
        "requests": [
          {
            "url": (self._get_ms_graph().generate_graph_url(
                      resource= self._get_ms_graph_resource(), 
                      resource_id= self._get_ms_graph_resource_id(), 
                      base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/range(address='A{group_data.get('start_index')}:{get_column_letter(len(group_data.get('data')[0]))}{group_data.get('end_index')}')?$select=id,values,columnCount")),
            "method": "PATCH",
            "id": f"{group_data.get('start_index')}_{group_data.get('end_index')}",
            "body": {
              "values" : group_data.get('data'),
            },
            "headers": {
                "Content-Type": "application/json"
            }
          } for group_data in update_data
        ]
      },
      method= "post"
    ))
      

  def __sync_data_update_data(self, sheet_key, update_data, *args, **kwargs):
    
    if len(update_data) < 1:
      return
    
    self.__close_session()
    update_data.sort(key=lambda ceil: ceil.get("index"))
    update_data_parsed = []
    
    batch_size = 15
    batch_count = 1
    # JSON batch requests are currently limited to 20 individual requests in addition to the following limitations:
    # https://learn.microsoft.com/en-us/graph/json-batching#batch-size-limitations
    while len(update_data) > 0:
      if batch_count >= 2000:
        sleep(5)

      row = update_data.pop(0)
      index_row_offset = 2

      if len(update_data_parsed) < 1 or ((update_data_parsed[-1]["end_index"] + 1) != (row.get("index") + index_row_offset)):
        update_data_parsed.append({
          "start_index": row.get("index") + index_row_offset,
          "end_index": row.get("index") + index_row_offset,
          "data": row.get("values")
        })
        continue

      update_data_parsed[-1]["end_index"] += 1
      update_data_parsed[-1]["data"].append(row.get("values")[0])

      if len(update_data_parsed) >= batch_size:
        try:
          self.__sync_data_update_data_process(
            sheet_key= sheet_key,
            update_data= update_data_parsed,
          )
          update_data_parsed.clear()
          batch_count += 1
        except:
          pass
    
    if len(update_data_parsed) > 0:
      try:
        self.__sync_data_update_data_process(
          sheet_key= sheet_key,
          update_data= update_data_parsed,
        )
        update_data_parsed.clear()
        batch_count += 1
      except:
        pass

      
    
  
  
  def __sync_data_add_data(self, sheet_key, insert_data, *args, **kwargs):
    
    data_len = len(insert_data)
    if data_len < 1:
      return
    
    batch_size = self._get_ms_graph().max_batch_size
    groups = self._get_number_of_groups_batchsize(total_items= data_len, batch_size= batch_size)
      

    
    if groups > 100:
      batch_size = data_len / Decimal(100)
      batch_size = int(floor(batch_size)) + 1
      groups = self._get_number_of_groups_batchsize(total_items= data_len, batch_size= batch_size)

    return_results = []
    try:
      for iteration in range(int(groups)):
        start_index = iteration * batch_size
        end_index = start_index + batch_size

        return_results.append(
          self._get_ms_graph().send_request(
            url = self._get_ms_graph().generate_graph_url(
              resource= self._get_ms_graph_resource(), 
              resource_id= self._get_ms_graph_resource_id(), 
              base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{self._get_worksheet_table_data()[sheet_key].get('id')}/rows"),
            session_config = {
              "type":"workbook",
              "graph_resource": self._get_ms_graph_resource(),
              "graph_resource_id": self._get_ms_graph_resource_id(),
              "drive_id": self.get_cmdb_file().get('id'),
              "persist_changes": True,
              "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
            },
            data = {
              "values": insert_data[start_index:(end_index if end_index < data_len else None)]
            },
            method= "post"
          )
        )

    except Exception as err:
      self.__close_session()
      raise err
    
    return return_results
  
  def __sync_data_get_existing_has_data(self, existing_data, *args, **kwargs):
    if existing_data is None:
      return False
    
    if existing_data.get("value") is None:
      return False
    
    if len(existing_data.get("value")) < 1:
      return False
    
    return True

  def __sync_data_get_existing(self, sheet_key, *args, **kwargs):
    
    try:
      existing_data = self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{self._get_worksheet_table_data()[sheet_key].get('id')}/rows"),
        session_config = {
          "type":"workbook",
          "graph_resource": self._get_ms_graph_resource(),
          "graph_resource_id": self._get_ms_graph_resource_id(),
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": False
        },
        method= "get"
      )
    except Exception as err:
      self.__close_session()
      raise err

    if not self.__sync_data_get_existing_has_data(existing_data= existing_data):
      return {}
    
    source_index = self.get_existing_columns_sorted_by_index()[sheet_key].index(self.get_cmdb_data_containers_columns_raw_byid_display()[sheet_key]["source"])
    cmdb_id_index = self.get_existing_columns_sorted_by_index()[sheet_key].index(self.get_cmdb_data_containers_columns_raw_byid_display()[sheet_key]["cmdb_id"])

    return {
      self.get_common().helper_type().string().set_case(
        string_value= data.get("values")[0][cmdb_id_index],
        case= "lower"
      ):data for data in existing_data.get("value") if self.get_common().helper_type().string().set_case(
        string_value= data.get("values")[0][source_index],
        case= "lower"
      ) == self.get_common().helper_type().string().set_case(
        string_value= self.get_cloud_client().get_provider(),
        case= "lower"
      )
    }


    




    
    
  
