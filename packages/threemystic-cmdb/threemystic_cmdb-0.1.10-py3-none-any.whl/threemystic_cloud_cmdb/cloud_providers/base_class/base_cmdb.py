from threemystic_common.base_class.base_provider import base
import asyncio, concurrent.futures
from abc import abstractmethod
from openpyxl import Workbook,worksheet
from polling2 import TimeoutException, poll as poll2
from random import Random

class cloud_cmdb_provider_base_cmdb(base):
  def __init__(self, *args, **kwargs):
    super().__init__(*args, **kwargs)
    
    self._set_max_process_pool(*args, **kwargs)
    self._set_max_thread_pool(*args, **kwargs)
    self._set_data_start(*args, **kwargs)
    self._set_cloud_cmdb(*args, **kwargs)
    self._set_client_name(*args, **kwargs)
    self.__set_cmdb_data_action(*args, **kwargs)
    self.__set_cmdb_cloud_share(*args, **kwargs)

  def _get_default_columns_cmdb_raw(self, *args, **kwargs):
    if hasattr(self, "_columns_cmdb_raw"):
      return self._columns_cmdb_raw
    self._columns_cmdb_raw = {
      "resource_group": {
        "id": "resource_group",
        "display": "ResourceGroup"
      },
      "region": {
        "id": "region",
        "display": "Region"
      },
      "environment": {
        "id": "environment",
        "display": "Environment"
      },
    }
    return self._get_default_columns_cmdb_raw()
  
  @abstractmethod
  def generate_resource_tags_csv(self, tags, seperator=",", tag_attribute_seperator=":", *args, **kwargs):
        pass
  @abstractmethod
  def _load_cmdb_general_data(self, *args, **kwargs):
    pass
  
  @abstractmethod
  def _load_cmdb_column_data(self, *args, **kwargs):
    pass
  
  @abstractmethod
  def _get_default_columns_raw(self, *args, **kwargs):
    pass
  
  @abstractmethod
  def _get_report_default_row(self, account, *args, **kwargs):
    pass
  
  @abstractmethod
  def _get_report_default_row_cmdb(self, account, *args, **kwargs):
    pass
  
  @abstractmethod
  def generate_tag_columns(self, account, resource, is_cmdb = False, *args, **kwargs):
    pass
  
  def get_account_environment(self, account = None, resource = None, *args, **kwargs):
    # for simplicity this is handeled on the data side
    if resource is not None and not self.get_common().helper_type().string().is_null_or_whitespace(string_value= resource.get("extra_environment")):
      return resource.get("extra_environment")
    
    if account is not None:
      return account.get("extra_environment")
    
    return "unknown"

  def get_default_columns(self, *args, **kwargs):
    return [
      column.get("display") for column in self._get_default_columns_raw(*args, **kwargs) if column.get("hidden") != True
    ]

  def get_item_data_value(self, item_data, value_key, *args, **kwargs):
    return self.get_common().helper_type().general().get_container_value(container= item_data, value_key= value_key)
  
  def save_excel_report_path(self, *args, **kwargs):
    if hasattr(self, "_excel_report_path_save" ):
      return self._excel_report_path_save

    report_name = f'{self.get_client_name()}-{self.get_common().helper_type().datetime().datetime_as_string(dt= self.get_data_start(), dt_format= "%Y%m%d")}.xlsx'
    self._excel_report_path_save = self.get_cloud_cmdb().get_cmdb_report_path().joinpath(f'{self._get_cloud_cmdb_raw().get_provider()}/{report_name}')
    if not self._excel_report_path_save.parent.exists():
      self._excel_report_path_save.parent.mkdir(parents= True)
    
    return self.save_excel_report_path()
  
  async def save_excel_report_only(self, close_connection = False, *args, **kwargs):
    if hasattr(self, "_saving_excel_report_data"):
      if self._saving_excel_report_data is True and not close_connection:
        return
      
      if self._saving_excel_report_data is True and not close_connection:
        poll2(
          lambda: self._saving_excel_report_data,
          ignore_exceptions=(Exception,),
          timeout= 240,
          step=0.1
        )
    
    self._saving_excel_report_data = True

    # When in write mode you can only save onces
    self._get_excel().save(self.save_excel_report_path())
    self._saving_excel_report_data = False

    if close_connection is True:
      self._excel_close()

  def _stop_is_processing_sheet_data_or_set_processing(self, sheet_key, *args, **kwargs):
    setattr(self, f"_is_processing_sheet_data_{sheet_key}", False)

  def _is_processing_sheet_data_or_set_processing(self, sheet_key, *args, **kwargs):
    if getattr(self, f"_is_processing_sheet_data_{sheet_key}", False) is True:
      return True
        
      
    setattr(self, f"_is_processing_sheet_data_{sheet_key}", True)
    return False

  async def save_report(self, *args, **kwargs):
    
    
    if len(self._get_excel().sheetnames) < 1:
      print(f'No Report Saved - No Data')
      return  
    print(f'Report saved at: {self.save_excel_report_path()}')
    await self.save_excel_report_only(close_connection= True)
    
  
  async def save_report_cmdb(self, *args, **kwargs):    
    if not self.get_cloud_cmdb().has_cloud_share_configured():
      return None
    
    print(f'Saving report in Shared CMDB')
    self.get_cmdb_cloud_share().save_data(
      report_data= self.get_cmdb_workbook(sheet_key= None)
    )
    

  def get_report_default_row(self, account, sheet_key, resource = None, is_cmdb = False,  *args, **kwargs):
    default_row = self._get_report_default_row(account= account) if is_cmdb == False else self._get_report_default_row_cmdb(account= account)

    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_resourcegroup") is True:
      if not is_cmdb:
        default_row.append(self.get_common().helper_type().string().join(",", resource.get("extra_resourcegroups")))
      else:
        default_row["resource_group"] = self.get_common().helper_type().string().join(",", resource.get("extra_resourcegroups"))

    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_environment") is True:
      if not is_cmdb:
        default_row.append(self.get_account_environment(account= account, resource= resource))
      else:
        default_row["environment"] = self.get_account_environment(account= account, resource= resource)
      

    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_region") is True:
      if not is_cmdb:
        default_row.append(self.get_cloud_client().get_azresource_location(resource= resource))
      else:
        default_row["region"] = self.get_cloud_client().get_azresource_location(resource= resource)
      
    
    return default_row

  async def __main_poolexecutor(self, *args, **kwargs):   
    with concurrent.futures.ThreadPoolExecutor(self.get_max_thread_pool()) as pool:
        return await self.main_process(
          pool= pool,
          **kwargs
        )

  async def main(self, pool= None, *args, **kwargs):   
    if pool == None:
      return await self.__main_poolexecutor(*args, **kwargs)
    
    return await self.main_process(
          pool= pool,
          **kwargs
        )
  
  async def main_process(self, pool, loop= None, *args, **kwargs):    
    
    print(f"Running: {self.get_client_name()} - {self.get_data_start()}")
    await self._generate_report_data(
      pool= pool,
      loop= loop,
      *args, **kwargs
    )
    await self.save_report()
    await self.save_report_cmdb()

  
  def _get_data_action_by_key(self, *args, **kwargs):
    return {
      action: self._get_cloud_cmdb_raw().get_cloud_data_client().get_data_action(action= action) for action in list(self.get_workbook_general_data().keys())
    }
  
  async def _generate_report_data(self, pool, loop= None, *args, **kwargs):  
    
    # await self._process_report_data(
    #   data= {
    #     sheet_key: await data.main(pool= pool, loop= loop,) for sheet_key, data in self._get_data_action_by_key().items()
    #   }
    # )
    for sheet_key, data in self._get_data_action_by_key().items():
      await data.main(
          run_params = {
            "send_account_data_lambda": {
              "handler": self._process_report_data,
              "sheet_key": sheet_key
            },
          },
          pool= pool,
          loop= loop)

  async def _process_report_data(self, data, *args, **kwargs):
    
    for sheet_key, main_report_data in data.items():
      while self._is_processing_sheet_data_or_set_processing(sheet_key= sheet_key):
        await asyncio.sleep(Random().random()*2)
      for _, report_data in main_report_data.items():
        if report_data is None:
          continue
        
        while len(report_data) > 0:
          report_data_item=report_data.pop(0)
          self.get_excel_workbook(sheet_key= sheet_key).append(
            self.get_report_default_row(
              sheet_key= sheet_key,
              account= report_data_item.get("extra_account"), 
              resource= report_data_item,  
              region= report_data_item.get("extra_region"), 
              resource_groups = report_data_item.get("extra_resourcegroups"),
              is_cmdb= False
              ) +
            [self.get_handler_column_data(column_data= column_data, is_cmdb= False)(report_data_item) for _, column_data in self.get_workbook_columns()[sheet_key].items()] +
            self.generate_tag_columns(
              account=report_data_item.get("extra_account"),
              resource= report_data_item,
              is_cmdb= False)
          )
          self.save_cmdb_workbook_item(sheet_key= sheet_key, account= report_data_item.get("extra_account"), report_data_item= report_data_item)
      
      self._stop_is_processing_sheet_data_or_set_processing(sheet_key= sheet_key)


      

  def _excel_close(self, *args, **kwargs):    
    self._get_excel().close()
    self.get_excel_workbook(unset= True)
    self._get_excel(unset= True)

  def _get_excel(self, unset = False, *args, **kwargs):    
    if hasattr(self, "_workbook_excel_main"):
      if unset:
        return delattr(self, "_workbook_excel_main")
      return self._workbook_excel_main
    
    self._workbook_excel_main = Workbook(write_only=True)
    
    while len(self._workbook_excel_main.sheetnames) > 0:
      self._workbook_excel_main.remove(self._workbook_excel_main[self._workbook_excel_main.sheetnames[0]])
    return self._get_excel()

  def get_cmdb_workbook(self, sheet_key= None, *args, **kwargs):
    if not self.get_cloud_cmdb().has_cloud_share_configured():
      return None
    
    if hasattr(self, "_workbook_cmdb_data"):
      if sheet_key is None:
        return self._workbook_cmdb_data
      
      if self._workbook_cmdb_data.get(sheet_key) is not None:
        return self._workbook_cmdb_data[sheet_key]
    
    if not hasattr(self, "_workbook_cmdb_data"):
      self._workbook_cmdb_data = {}

    self._workbook_cmdb_data[sheet_key] = []
    return self.get_cmdb_workbook(sheet_key= sheet_key, *args, **kwargs)
  
  def get_ishidden_column_data(self, column_data, is_cmdb = False, *args, **kwargs):    
    if not is_cmdb or column_data.get("cmdb") is None:
      return column_data.get("hidden") is True
    
    if column_data.get("cmdb").get("hidden") is None:
      return column_data.get("hidden") is True
    
    return column_data.get("cmdb").get("hidden") is True
  
  def get_handler_column_data(self, column_data, is_cmdb = False, *args, **kwargs):    
    if not is_cmdb or column_data.get("cmdb") is None:
      return column_data["handler"]
    
    if column_data.get("cmdb").get("handler") is None:
      return column_data["handler"]
    
    return column_data.get("cmdb").get("handler")

  def get_display_column_data(self, column_data, is_cmdb = False, *args, **kwargs):
    if not is_cmdb or column_data.get("cmdb") is None:
      return str(column_data.get("display"))
    
    if self.get_common().helper_type().string().is_null_or_whitespace(string_value= column_data.get("cmdb").get("display")):
      return str(column_data["display"])
    
    return str(column_data.get("cmdb").get("display"))
  
  def save_cmdb_workbook_item(self, sheet_key, account, report_data_item, *args, **kwargs):
    if not self.get_cloud_cmdb().has_cloud_share_configured():
      return None

    self.get_cmdb_workbook(sheet_key= sheet_key).append({
      "data":
        self.get_common().helper_type().dictionary().merge_dictionary([
          {},
          self.get_report_default_row(
            sheet_key= sheet_key,
            account= account, 
            resource= report_data_item,  
            region= report_data_item.get("extra_region"), 
            resource_groups = report_data_item.get("extra_resourcegroups"),
            is_cmdb = True),
          {
            # This needs to be updated to ID
            column_data.get("id"):self.get_handler_column_data(column_data= column_data, is_cmdb= True)(report_data_item)
            for column_data in (self.get_workbook_columns()[sheet_key].values())},
          self.generate_tag_columns(
            account= account, 
            resource= report_data_item,
            is_cmdb= True)
        ])
      }
    )
    self.get_cmdb_workbook(sheet_key= sheet_key)[-1]["raw_data"] = report_data_item
    

  def get_excel_workbook(self, sheet_key = None, unset = False, *args, **kwargs):    
    if hasattr(self, "_workbook_excel_data"):
      if unset:
        return delattr(self, "_workbook_excel_data")
      
      if sheet_key is not None and self._workbook_excel_data.get(sheet_key) is not None:
        return self._workbook_excel_data[sheet_key]
      
      if sheet_key is None:
        return self._workbook_excel_data
    
    if not hasattr(self, "_workbook_excel_data"):
      self._workbook_excel_data = {}

    self._workbook_excel_data[sheet_key] = self._get_excel().create_sheet(
      title= self.get_workbook_general_data(sheet_key= sheet_key)["display"]
    )    
    
    self._workbook_excel_data[sheet_key].append(
      self.get_default_report_columns(sheet_key= sheet_key, *args, **kwargs) +
      [
        self.get_workbook_column_header_display(info_column= column) for _, column in self.get_workbook_columns()[sheet_key].items()
      ] +
      self.get_tag_report_columns(sheet_key= sheet_key, is_cmdb= False, *args, **kwargs)
    )
    return self.get_excel_workbook(sheet_key= sheet_key, *args, **kwargs)

  def get_cmdb_data_action(self, *args, **kwargs):
    return self.__cmdb_data_action
  
  def __set_cmdb_data_action(self, data_action, *args, **kwargs):
    self.__cmdb_data_action = data_action

  def get_cmdb_cloud_share(self, *args, **kwargs):
    return self.__cmdb_connector
  
  def __set_cmdb_cloud_share(self,*args, **kwargs):

    from threemystic_cloud_cmdb.cloud_providers.general.cmdb_connector.auto_cmdb_connector import cloud_cmdb_general_cmdb_connector_auto as cmdb_connector
    self.__cmdb_connector = cmdb_connector(
      common= self.get_common(),
      logger= self.get_common().get_logger(),
      cloud_client= self.get_cloud_client(),
      data_containers= self.get_workbook_general_data(),
      container_columns= self.generate_workbook_columns_data_cmdb(),
    
    ).get_connector()

  def get_data_start(self, *args, **kwargs):
    return self.__data_start
  
  def _set_data_start(self, *args, **kwargs):
    self.__data_start = self.get_common().helper_type().datetime().get()

  def get_cloud_cmdb(self, *args, **kwargs):
    return self._get_cloud_cmdb_raw().get_cloud_cmdb() 
  
  def get_cloud_client(self, *args, **kwargs):
    return self._get_cloud_cmdb_raw().get_cloud_data_client().get_cloud_client()
  
  def _get_cloud_cmdb_raw(self, *args, **kwargs):
    return self.__cloud_cmdb
  
  def _set_cloud_cmdb(self, cloud_cmdb, *args, **kwargs):
    self.__cloud_cmdb = cloud_cmdb

  def get_client_name(self, *args, **kwargs):
    return self._client_name
  
  def _set_client_name(self, data_action, *args, **kwargs):
    self._client_name = f"{self.get_provider()}-{data_action}-cmdb"

  def get_max_process_pool(self, *args, **kwargs):
    return self._max_process_pool
  
  def _set_max_process_pool(self, max_thread_pool = 5, *args, **kwargs):
    self._max_process_pool = max_thread_pool

  def get_max_thread_pool(self, *args, **kwargs):
    return self._max_thread_pool
  
  def _set_max_thread_pool(self, max_thread_pool = 35, *args, **kwargs):
    self._max_thread_pool = max_thread_pool

  async def _pre_load_main_process(self, *args, **kwargs):
    pass
  
  def get_workbook_general_data(self, sheet_key = None, *args, **kwargs):
    if hasattr(self, "_workbook_general_data"):
      if self.get_common().helper_type().string().is_null_or_whitespace(string_value= sheet_key):
        return self._workbook_general_data
      return self._workbook_general_data.get(sheet_key)
        
    
    self._set_workbook_general_data()
    return self.get_workbook_general_data(sheet_key= sheet_key, *args, **kwargs)
  
  def _set_workbook_general_data(self, *args, **kwargs):
    self._workbook_general_data = self._generate_workbook_data()

  def _load_cmdb_general_default_data(self, *args, **kwargs):
    return {
      "include_region": True,
      "include_environment": True,
      "include_resourcegroup": True,
      "include_requiredtags": True,
      "cmdb_connector":{
        "include_delete_column": True,
        "include_empty_column": True,
      }
    }
  
  def _load_cmdb_general_default_column_data(self, *args, **kwargs):
    return {
      "display": "Data",
      "handler": lambda item: item
    }


  def _generate_workbook_data(self, *args, **kwargs):
    return_data = {}
    for key, item in self._load_cmdb_general_data().items():
      return_data[key] = self.get_common().helper_type().dictionary().merge_dictionary([
        {},
        self._load_cmdb_general_default_data(),
        item
    ])
    return return_data
  
  
  def get_workbook_columns(self, *args, **kwargs):
    if hasattr(self, "_workbook_column_data"):
      return self._workbook_column_data
    
    self._set_workbook_columns(*args, **kwargs)
    return self.get_workbook_columns(*args, **kwargs)
  
  def _set_workbook_columns(self, *args, **kwargs):
    self._workbook_column_data = self._generate_workbook_columns_data()

  def generate_workbook_columns_data_cmdb(self, *args, **kwargs):
    return_column_cmdb_data = {}
    
    for sheet_key, item in self._load_cmdb_column_data().items():
      return_column_cmdb_data[sheet_key]= (
        self.get_default_report_columns(sheet_key= sheet_key, is_cmdb= True) +
        [self.get_common().helper_type().dictionary().merge_dictionary([
          {},
          {"id": item_key},
          item_data
        ]) for item_key, item_data in item.items()] +
        [self.get_common().helper_type().dictionary().merge_dictionary([
          {},
          {"id": item_key},
          item_data
        ]) for item_key, item_data in self.get_tag_report_columns(sheet_key= sheet_key, is_cmdb= True).items()]
      )

    return return_column_cmdb_data

    
    
  def _generate_workbook_columns_data(self, *args, **kwargs):
    return_data = {}
    for key, item in self._load_cmdb_column_data().items():
      
      return_data[key] = {
        item_key:self.get_common().helper_type().dictionary().merge_dictionary([
          {},
          self._load_cmdb_general_default_column_data(),
          item_data,
          {"id": item_key}
        ]) for item_key, item_data in item.items()
      }
        
    return return_data
  
  def get_workbook_column_header_display(self, info_column, *args, **kwargs):
    
    if self.get_common().helper_type().general().is_type(info_column["display"], str):
      return info_column["display"]

    
    raise self.get_common().exception().exception(
        exception_type = "argument"
      ).not_implemented(
        logger = self.get_common().get_logger(),
        name = "info_column",
        message = f"Unknown info_column: {self.get_common().helper_json().dumps(data= info_column)}"
      )
  
  def get_default_report_columns(self, sheet_key, is_cmdb = False, *args, **kwargs):

    default_columns = []
    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_region") is True:      
      default_columns.append(self._get_default_columns_cmdb_raw()["region"].get("display") if not is_cmdb else self._get_default_columns_cmdb_raw()["region"])
    
    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_resourcegroup") is True:
      default_columns.append(self._get_default_columns_cmdb_raw()["resource_group"].get("display") if not is_cmdb else self._get_default_columns_cmdb_raw()["resource_group"])
      
    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_environment") is True:
      default_columns.append(self._get_default_columns_cmdb_raw()["environment"].get("display") if not is_cmdb else self._get_default_columns_cmdb_raw()["environment"])

    return (
      (self.get_default_columns() if not is_cmdb else self._get_default_columns_raw(*args, **kwargs)) +
      default_columns
    )
  
  def get_tag_report_columns(self, sheet_key, is_cmdb = False, *args, **kwargs):
    if self.get_workbook_general_data(sheet_key= sheet_key).get("include_requiredtags") is True:
      if not is_cmdb:
        return []
      
      return {}

    return [] if not is_cmdb else {}
    # the plan is the required tags will be in the config
    # return {
    #   # f'{prefix}{tag}' for tag in self.required_tag_names()
    # }
  
  # def required_tag_names(cls):
  #   return {
  #     "Name": None,
  #     # "ManagedBy": {
  #     #   "basic": ["Supported By", "Application Contact"],
  #     #   "custom": {
  #     #     "ManagedByTerraform": lambda item: "terraform" if cls.is_true(item) else None
  #     #   }
  #     # },
  #     "TechOwner": None,
  #     "TechContact": None,
  #     "Environment": None,
  #     "CodeRepo": None,
  #     "Product": None
  #   }